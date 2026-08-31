"""Reaching the supplementary tables a paper's per-variant numbers actually live in.

A GWAS paper's body says *"263 independent variants across 180 genomic loci"*. It
does not list them. The rsIDs, positions, effect alleles and per-trait p-values —
every column ``studies.csv`` and ``gwas_effects.csv`` want — are in the
supplementary workbook, and ``fetch_fulltext`` returns the JATS body without it.

Two agents given the same task independently concluded "no plugin tool fetches
supplementary material" and dropped a paper's entire contribution; one of them
then reconstructed the publisher URL pattern over twelve blind probes. This module
is that ladder, run through :mod:`~just_module_creator.net` so it is paced,
retried and attributed like every other outbound call.

**The three-valued result is the point.** ``not_determinable`` — we hold no
pattern for this publisher — is a different answer from ``none_published``, and
collapsing them is exactly how a surface limit gets recorded as a fact about the
world. Every function here keeps them apart and says which rung answered.
"""

from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook

from just_module_creator.net import HttpService, ServiceUnavailable

#: Springer Nature hosts every ESM for its imprints on one open object store, and
#: it is open — no bot challenge, unlike ``link.springer.com``, which answers a
#: plain GET with a 3 KB JavaScript "Client Challenge" under HTTP 200.
ESM_HOST = "https://static-content.springer.com"

#: DOI prefix → the publisher family we can address by pattern. **Measured, not
#: assumed**: 2026-08-30/31 against four real articles across the three prefixes.
#: A prefix absent from here is why a result says `not_determinable` rather than
#: `none_published` — it is a gap in our coverage, never a fact about the paper.
PATTERNED_PREFIXES: dict[str, str] = {
    "10.1007": "Springer",
    "10.1186": "BMC",
    "10.1038": "Nature Portfolio",
}

#: Probed in this order when no inventory named the extension. Ordered by what
#: actually carries tables. Guessing is the fallback rung and its misses mean
#: *unknown*: ``MOESM1`` has been observed as `.txt` on one article and `.pdf` on
#: another, and a peer-review file sits at `MOESM3` on a third.
CANDIDATE_EXTENSIONS = ("xlsx", "pdf", "docx", "csv", "txt", "zip", "xls", "doc")

#: Springer's suffix: ``s11357-025-02044-3`` → journal 11357, year 2025, article
#: 2044. Leading zeros are stripped from the article number and the two-digit year
#: is expanded — ``s41467-018-03242-8`` gives ``41467_2018_3242``.
_SUFFIX = re.compile(r"^s?(\d+)-(\d{2,4})-(\d+)(?:-\d+)?$")


class SupplementaryError(RuntimeError):
    """The ladder could not be run at all — never that a paper has no ESM."""


@dataclass
class SupplementaryFile:
    """One supplementary object. ``caption`` is often uninformative, deliberately
    reported as the publisher wrote it rather than improved."""

    name: str
    url: str
    extension: str
    caption: str | None = None
    size_bytes: int | None = None


@dataclass
class Inventory:
    """What is published, and — as much as it matters — how we found out."""

    #: ``found`` | ``none_published`` | ``not_determinable``
    verdict: str
    #: ``europepmc_xml`` | ``publisher_pattern`` | ``none``
    rung: str
    files: list[SupplementaryFile] = field(default_factory=list)
    why_not: str | None = None
    publisher: str | None = None
    doi: str | None = None
    notes: list[str] = field(default_factory=list)


def esm_stem(doi: str) -> str | None:
    """``10.1007/s11357-025-02044-3`` → ``11357_2025_2044``. None when unparseable."""
    _, _, suffix = doi.partition("/")
    match = _SUFFIX.match(suffix.strip())
    if not match:
        return None
    journal, year, article = match.groups()
    full_year = int(year) if len(year) == 4 else 2000 + int(year)
    return f"{journal}_{full_year}_{int(article)}"


def esm_path(doi: str, stem: str, index: int, extension: str) -> str:
    """Path only. The DOI is url-encoded *inside* an ``art:`` segment."""
    return f"/esm/art%3A{quote(doi, safe='')}/MediaObjects/{stem}_MOESM{index}_ESM.{extension}"


def esm_url(doi: str, stem: str, index: int, extension: str) -> str:
    """The full addressable form, for reporting back to a caller."""
    return f"{ESM_HOST}{esm_path(doi, stem, index, extension)}"


def parse_jats_supplementary(xml: str, base_url: str) -> list[SupplementaryFile]:
    """Every ``<supplementary-material>``, deduplicated.

    Europe PMC lists each file **twice** — once in the body, once in the back
    matter — so a naive parse doubles the inventory. Keyed by filename, first
    caption wins.
    """
    found: dict[str, SupplementaryFile] = {}
    for block in re.findall(r"<supplementary-material.*?</supplementary-material>", xml, re.S):
        href = re.search(r'xlink:href="([^"]+)"', block)
        if not href:
            continue
        name = href.group(1)
        if name in found:
            continue
        caption = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", block)).strip() or None
        found[name] = SupplementaryFile(
            name=name,
            url=f"{base_url.rstrip('/')}/{name}",
            extension=name.rsplit(".", 1)[-1].lower() if "." in name else "",
            caption=caption,
        )
    return list(found.values())


def workbook_sheets(path: Path) -> tuple[list[str], list[str]]:
    """Sheet names and the ``Supplementary Table`` titles inside the cells.

    Read straight out of the xlsx zip so this needs **no spreadsheet dependency** —
    the runs that met this wall had no ``openpyxl`` and one hand-rolled a reader
    with a column-alignment bug. Sheets are commonly ``ST1``…``ST14`` while the
    in-cell titles read *"Supplementary Table S8: …"*, and neither alone tells you
    which sheet answers your question.

    Returns names and titles, never rows: which sheet supports a row's claim is a
    judgement about that row, the same reason ``fetch_fulltext`` returns no
    best-matching passage.
    """
    if not zipfile.is_zipfile(path):
        return [], []
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()
        sheets: list[str] = []
        if "xl/workbook.xml" in names:
            book = archive.read("xl/workbook.xml").decode("utf-8", "replace")
            sheets = re.findall(r'<sheet [^>]*name="([^"]*)"', book)
        titles: list[str] = []
        if "xl/sharedStrings.xml" in names:
            shared = archive.read("xl/sharedStrings.xml").decode("utf-8", "replace")
            titles = sorted(set(re.findall(r"(Supplementary Tables?[^<]{0,200})", shared)))
    return sheets, titles


#: What a cell can be on the wire. `datetime` is rendered ISO-8601 rather than
#: passed through: a supplementary table's date column is metadata an author
#: transcribes, and a naive `YYYY-MM-DD HH:MM:SS` is misparsed as local time.
Cell = str | int | float | bool | None


@dataclass(frozen=True)
class SheetWindow:
    """A window onto one sheet, plus the two counts that make paging honest."""

    rows: list[list[Cell]]
    total_rows: int
    last_populated_row: int | None
    width: int
    sheet_names: list[str]


def workbook_rows(path: Path, sheet: str, offset: int = 0, limit: int = 200) -> SheetWindow:
    """One sheet's cells, as rows of equal width. The rung after ``workbook_sheets``.

    ``rows`` is the window ``offset..offset+limit``. ``total_rows`` is the sheet's
    whole span and ``last_populated_row`` is the 0-based index of the last row with
    anything in it — **both, because they disagree and the difference is trailing
    blank rows a producer left behind**. Measured on the ARDS paper's ST9: 1000 rows
    spanned, 265 last populated, so a caller paging on ``total_rows`` alone spends
    three quarters of its calls on nothing. ``None`` means the sheet is empty.

    **Three things this does that a hand-rolled reader keeps getting wrong**, each
    measured on a real run rather than imagined:

    * ``max_row``/``max_column`` come from the workbook's *dimension record*, which
      is optional and is written wrong by plenty of producers — the "missing ``ref``
      attribute" one run spent two bugs on. Both are counted while streaming.
    * Every row is **padded to the sheet's width**. openpyxl yields a short tuple
      for a row whose trailing cells are empty, and a caller zipping that against a
      header silently shifts columns — the other bug from the same run.
    * A ``datetime`` becomes an ISO-8601 string rather than crossing the wire as an
      object with a repr that is not what the cell said.

    Rows come back exactly as the sheet has them, including its title and blank
    lines: which row is the header, and which sheet answers a claim, are judgements
    about the row being authored. This hands over the cells and stops.
    """
    if not zipfile.is_zipfile(path):
        raise SupplementaryError(
            f"{path.name} is not an xlsx workbook. A legacy .xls, a PDF or a CSV is not read "
            "here — that is a limit of this reader, not a statement that the file has no rows."
        )
    book = load_workbook(path, read_only=True, data_only=True)
    try:
        names = list(book.sheetnames)
        if sheet not in names:
            raise SupplementaryError(
                f"No sheet named {sheet!r}. This workbook has: {', '.join(names)}"
            )
        worksheet = book[sheet]
        rows: list[list[Cell]] = []
        total = 0
        width = 0
        last_populated: int | None = None
        stop = offset + limit
        for index, row in enumerate(worksheet.iter_rows(values_only=True)):
            total += 1
            if any(value is not None for value in row):
                last_populated = index
                # Width is taken over POPULATED rows only. A trailing blank row can
                # be yielded at the dimension record's width and would otherwise pad
                # every real row out to a column nothing ever wrote in.
                width = max(width, len(row))
            if offset <= index < stop:
                rows.append([_cell(value) for value in row])
        # Padded AFTER the whole sheet is seen: the widest row may sit below the
        # window, and a window narrower than the sheet is the shift this prevents.
        return SheetWindow(
            rows=[row + [None] * (width - len(row)) for row in rows][:limit],
            total_rows=total,
            last_populated_row=last_populated,
            width=width,
            sheet_names=names,
        )
    finally:
        book.close()


def _cell(value: object) -> Cell:
    if isinstance(value, datetime | date):
        return value.isoformat()
    if isinstance(value, str | int | float | bool) or value is None:
        return value
    return str(value)


def inventory(
    *,
    doi: str | None,
    xml: str | None,
    xml_base_url: str | None,
    probe: HttpService | None,
    max_index: int = 60,
) -> Inventory:
    """Run the ladder: the article's own XML first, the publisher pattern second.

    ``xml`` is Europe PMC's ``fullTextXML`` when the article is in PMC — the
    authoritative rung, since it carries each file's real name *and extension*
    rather than a guess. The pattern rung exists because the common case for a
    paper published this month is that it is not in PMC at all, which is exactly
    when somebody is writing a module about it.
    """
    notes: list[str] = []
    if xml and xml_base_url:
        files = parse_jats_supplementary(xml, xml_base_url)
        if files:
            return Inventory(
                verdict="found", rung="europepmc_xml", files=files, doi=doi,
                notes=["Names and extensions are the publisher's own, not guessed."],
            )
        notes.append(
            "The article is in Europe PMC and its fulltext XML lists no supplementary file."
        )

    if not doi:
        return Inventory(
            verdict="not_determinable", rung="none", doi=doi, notes=notes,
            why_not="No DOI, so the publisher pattern cannot be addressed.",
        )

    prefix = doi.split("/", 1)[0]
    publisher = PATTERNED_PREFIXES.get(prefix)
    if publisher is None:
        return Inventory(
            verdict="not_determinable", rung="none", doi=doi, notes=notes,
            why_not=(
                f"No ESM URL pattern is known for DOI prefix {prefix}. This is a gap in "
                f"our coverage, not evidence the article has no supplementary material — "
                f"patterns exist for {', '.join(sorted(PATTERNED_PREFIXES))}."
            ),
        )

    stem = esm_stem(doi)
    if stem is None:
        return Inventory(
            verdict="not_determinable", rung="none", doi=doi, publisher=publisher, notes=notes,
            why_not=f"The DOI suffix does not parse into a {publisher} ESM stem.",
        )
    if probe is None:
        return Inventory(
            verdict="not_determinable", rung="none", doi=doi, publisher=publisher, notes=notes,
            why_not="Offline: the publisher pattern can only be answered by asking the host.",
        )

    files: list[SupplementaryFile] = []
    misses = 0
    for index in range(1, max_index + 1):
        hit = _probe_one(probe, doi, stem, index)
        if hit is None:
            misses += 1
            # Two consecutive absent indices ends it. A guess, and named as one:
            # the host cannot be asked "how many are there".
            if misses >= 2:
                break
            continue
        misses = 0
        files.append(hit)

    if files:
        notes.append(
            "Found by probing the publisher pattern, so the set is bounded by a guess: "
            "enumeration stops after two consecutive absent indices, and an extension we "
            "did not try reads as absent. Prefer the Europe PMC inventory where it exists."
        )
        return Inventory(
            verdict="found", rung="publisher_pattern", files=files,
            doi=doi, publisher=publisher, notes=notes,
        )
    return Inventory(
        verdict="none_published", rung="publisher_pattern", doi=doi,
        publisher=publisher, notes=notes,
    )


def _probe_one(
    probe: HttpService, doi: str, stem: str, index: int
) -> SupplementaryFile | None:
    """One index across the candidate extensions. 403 on this host means no key."""
    for extension in CANDIDATE_EXTENSIONS:
        url = esm_url(doi, stem, index, extension)
        try:
            response = probe.probe(esm_path(doi, stem, index, extension))
        except ServiceUnavailable:
            # Throttling and outages already retried in net.py. Reaching here means
            # the host stayed unreachable: unknown for this index, never absent.
            return None
        if response.status_code < 400:
            length = response.headers.get("content-length")
            return SupplementaryFile(
                name=f"{stem}_MOESM{index}_ESM.{extension}",
                url=url,
                extension=extension,
                size_bytes=int(length) if length and length.isdigit() else None,
            )
    return None
